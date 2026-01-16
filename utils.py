"""
Utility functions and service layer for XML-Excel conversion
Contains ALL conversion logic - NO dependency on main.py
"""

import os
import sys
from pathlib import Path
from typing import Optional
import logging
import xml.etree.ElementTree as ET
from xml.dom import minidom
from collections import defaultdict

try:
    import openpyxl
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    raise ImportError("openpyxl is required. Install it with: pip install openpyxl")

logger = logging.getLogger(__name__)


# =============================================================================
# CUSTOM EXCEPTIONS
# =============================================================================

class ConversionError(Exception):
    """Base exception for conversion errors"""
    pass


class ValidationError(ConversionError):
    """Exception for validation errors"""
    pass


class FileProcessingError(ConversionError):
    """Exception for file processing errors"""
    pass


# =============================================================================
# XML TO EXCEL CONVERSION FUNCTIONS
# =============================================================================

def analyze_xml_structure(root):
    """Analyze XML structure to determine if multiple sheets are needed."""
    children = list(root)
    
    if not children:
        return {"Sheet1": [root]}
    
    tag_groups = defaultdict(list)
    for child in children:
        tag_groups[child.tag].append(child)
    
    if len(tag_groups) == 1:
        tag_name = list(tag_groups.keys())[0]
        return {tag_name: tag_groups[tag_name]}
    
    sheets = {}
    for tag, elements in tag_groups.items():
        sheets[tag] = elements
    
    return sheets


def extract_element_data(element):
    """Extract data from an XML element into a flat dictionary."""
    data = {}
    
    # Add attributes with @ prefix
    for attr, value in element.attrib.items():
        data[f"@{attr}"] = value
    
    # Add child elements
    for child in element:
        child_text = child.text.strip() if child.text else ""
        
        if len(child) > 0:
            nested_data = extract_element_data(child)
            for key, value in nested_data.items():
                data[f"{child.tag}.{key}"] = value
        else:
            if child.tag in data:
                i = 2
                while f"{child.tag}_{i}" in data:
                    i += 1
                data[f"{child.tag}_{i}"] = child_text
            else:
                data[child.tag] = child_text
            
            for attr, value in child.attrib.items():
                data[f"{child.tag}.@{attr}"] = value
    
    if element.text and element.text.strip() and len(element) == 0:
        data["_text"] = element.text.strip()
    
    return data


def get_all_columns(elements):
    """Get all unique column names from a list of elements."""
    columns = []
    seen = set()
    
    for element in elements:
        data = extract_element_data(element)
        for key in data.keys():
            if key not in seen:
                columns.append(key)
                seen.add(key)
    
    return columns


def style_header(ws, num_columns):
    """Apply styling to the header row."""
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for col in range(1, num_columns + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border


def auto_adjust_column_width(ws):
    """Automatically adjust column widths based on content."""
    for column_cells in ws.columns:
        max_length = 0
        column = column_cells[0].column_letter
        
        for cell in column_cells:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = max(adjusted_width, 10)


def xml_to_excel_internal(xml_path, output_path=None):
    """Convert XML to Excel - internal implementation."""
    try:
        tree = ET.parse(xml_path)
        root = tree.getroot()
    except ET.ParseError as e:
        raise ValueError(f"Invalid XML file: {e}")
    except FileNotFoundError:
        raise FileNotFoundError(f"XML file not found: {xml_path}")
    
    sheets_data = analyze_xml_structure(root)
    
    wb = Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)
    
    for sheet_name, elements in sheets_data.items():
        safe_sheet_name = sheet_name[:31]
        for char in [':', '/', '\\', '?', '*', '[', ']']:
            safe_sheet_name = safe_sheet_name.replace(char, "_")
        
        ws = wb.create_sheet(title=safe_sheet_name)
        columns = get_all_columns(elements)
        
        if not columns:
            ws.cell(row=1, column=1, value="No data found")
            continue
        
        for col_idx, col_name in enumerate(columns, 1):
            ws.cell(row=1, column=col_idx, value=col_name)
        
        style_header(ws, len(columns))
        
        for row_idx, element in enumerate(elements, 2):
            data = extract_element_data(element)
            for col_idx, col_name in enumerate(columns, 1):
                value = data.get(col_name, "")
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        auto_adjust_column_width(ws)
        ws.freeze_panes = "A2"
    
    if output_path is None:
        output_path = Path(xml_path).with_suffix(".xlsx")
    
    wb.save(output_path)
    return str(output_path)


# =============================================================================
# EXCEL TO XML CONVERSION FUNCTIONS
# =============================================================================

def sanitize_tag_name(name, preserve_special=True):
    """Sanitize a string to be a valid XML tag name.
    
    Standard approach: removes spaces, preserves hyphens and periods.
    
    Args:
        name: The string to sanitize
        preserve_special: If True, preserves hyphens and periods (standard approach)
    """
    if not name:
        return "element"
    
    name = str(name).strip()
    
    # Handle pure numeric values (like column headers that are just numbers)
    if name.replace('.', '').replace('-', '').isdigit():
        return f"Column{name.replace('.', '').replace('-', '')}"
    
    # Standard XML approach: remove spaces, preserve hyphens and periods
    sanitized = ""
    for i, char in enumerate(name):
        if i == 0:
            # First character must be letter or underscore
            if char.isalpha() or char == '_':
                sanitized += char
            elif char.isdigit():
                # Start with underscore if first char is digit
                sanitized += "_" + char
            else:
                sanitized += ""  # Skip invalid first characters
        else:
            # Subsequent characters can include letters, digits, hyphens, underscores, periods
            if char.isalnum() or char in ['_', '-', '.']:
                sanitized += char
            elif char == ' ':
                # Remove spaces (standard approach - concatenate words)
                pass  # Skip spaces entirely
            else:
                sanitized += ""  # Remove other special characters
    
    if not sanitized or not (sanitized[0].isalpha() or sanitized[0] == '_'):
        sanitized = "element"
    
    if sanitized.lower().startswith("xml"):
        sanitized = "_" + sanitized
    
    return sanitized


def parse_column_structure(columns):
    """Parse column names to understand nested structure."""
    parsed = []
    for col in columns:
        col = str(col)
        if col.startswith("@"):
            parsed.append((col[1:], True, []))
        elif ".@" in col:
            parts = col.split(".@")
            path = parts[0].split(".")
            attr_name = parts[1]
            parsed.append((attr_name, True, path))
        elif "." in col:
            parts = col.split(".")
            parsed.append((parts[-1], False, parts[:-1]))
        else:
            parsed.append((col, False, []))
    
    return parsed


def row_to_xml_element(row_data, columns, element_tag):
    """Convert a row of data to an XML element."""
    element = ET.Element(sanitize_tag_name(element_tag))
    nested_elements = {}
    
    parsed_columns = parse_column_structure(columns)
    
    for (col_name, is_attr, nested_path), orig_col in zip(parsed_columns, columns):
        value = row_data.get(orig_col, "")
        if value is None or value == "":
            continue
        
        value = str(value)
        
        if not nested_path:
            if is_attr:
                element.set(sanitize_tag_name(col_name), value)
            else:
                child = ET.SubElement(element, sanitize_tag_name(col_name))
                child.text = value
        else:
            path_key = ".".join(nested_path)
            
            if path_key not in nested_elements:
                current = element
                for part in nested_path:
                    found = current.find(sanitize_tag_name(part))
                    if found is None:
                        found = ET.SubElement(current, sanitize_tag_name(part))
                    current = found
                nested_elements[path_key] = current
            
            target = nested_elements[path_key]
            
            if is_attr:
                target.set(sanitize_tag_name(col_name), value)
            else:
                child = ET.SubElement(target, sanitize_tag_name(col_name))
                child.text = value
    
    return element


def prettify_xml(element):
    """Return a pretty-printed XML string."""
    rough_string = ET.tostring(element, encoding='unicode')
    reparsed = minidom.parseString(rough_string)
    return reparsed.toprettyxml(indent="    ")


def detect_regions(ws):
    """Detect data regions in a worksheet.
    
    Standard approach: treat the entire used range as a single region.
    This ensures all rows are included, even those with empty cells.
    """
    max_row = ws.max_row
    max_col = ws.max_column
    
    # If sheet is empty, return empty list
    if max_row is None or max_col is None or max_row == 0 or max_col == 0:
        return []
    
    # Find the actual first row/column with data (header row)
    first_row = 1
    first_col = 1
    
    # Find first non-empty cell to start the region
    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            if ws.cell(row=r, column=c).value is not None:
                first_row = r
                first_col = c
                break
        else:
            continue
        break
    
    # Scan for actual last row with data
    real_max_row = first_row
    for r in range(max_row, first_row, -1):
        is_empty = True
        for c in range(1, max_col + 1):
            if ws.cell(row=r, column=c).value is not None:
                is_empty = False
                break
        if not is_empty:
            real_max_row = r
            break
            
    # Scan for actual last col with data
    real_max_col = first_col
    for c in range(max_col, first_col, -1):
        is_empty = True
        for r in range(1, real_max_row + 1):
            if ws.cell(row=r, column=c).value is not None:
                is_empty = False
                break
        if not is_empty:
            real_max_col = c
            break

    # Return the real used range
    return [(first_row, real_max_row, first_col, real_max_col)]


def excel_to_xml_internal(excel_path, output_path=None, root_tag=None):
    """Convert Excel to XML using standard flat structure.
    
    Produces XML matching standard converter output:
    - Root element is 'root' (or custom if specified)
    - Each row becomes an element named after the sheet (with appropriate transformation)
    - All columns are included, empty cells become self-closing tags
    - Structure is flat (no Table/Item wrappers)
    """
    try:
        wb = load_workbook(excel_path, data_only=True)
    except FileNotFoundError:
        raise FileNotFoundError(f"Excel file not found: {excel_path}")
    except Exception as e:
        raise ValueError(f"Invalid Excel file: {e}")
    
    # Use 'root' as default root tag (standard approach)
    if root_tag is None:
        root_tag = "root"
    
    root = ET.Element(root_tag)
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        # Detect data regions in the sheet
        regions = detect_regions(ws)
        
        for region_idx, (min_r, max_r, min_c, max_c) in enumerate(regions):
            # Get headers from first row of region
            # Standard approach: use cell value as tag name if it's text,
            # use positional Column{N} naming if value is numeric or empty
            headers = []
            for c in range(min_c, max_c + 1):
                val = ws.cell(row=min_r, column=c).value
                col_position = c - min_c + 1  # 1-based position within region
                
                if val is not None:
                    str_val = str(val).strip()
                    # Check if value is purely numeric (not a valid header name)
                    # Standard converters use Column{N} for numeric headers
                    try:
                        float(str_val)
                        # It's numeric, use positional column name
                        headers.append((c, f"Column{col_position}"))
                    except ValueError:
                        # It's text, use as header name
                        headers.append((c, str_val))
                else:
                    # Empty cell, use positional column name
                    headers.append((c, f"Column{col_position}"))
            
            if not headers:
                continue
            
            # Determine row element tag name from sheet name
            # Transform sheet name for XML compatibility
            row_tag = sanitize_tag_name(sheet_name)
            
            # Process each data row (skip header row)
            for r in range(min_r + 1, max_r + 1):
                row_element = ET.SubElement(root, row_tag)
                
                for col_idx, header_name in headers:
                    cell_value = ws.cell(row=r, column=col_idx).value
                    
                    # Sanitize header for tag name
                    tag_name = sanitize_tag_name(header_name)
                    
                    child = ET.SubElement(row_element, tag_name)
                    
                    if cell_value is not None:
                        # Convert value to string, handle special characters
                        str_value = str(cell_value)
                        child.text = str_value
                    # Empty cells become self-closing tags (child.text remains None)
    
    if output_path is None:
        output_path = Path(excel_path).with_suffix(".xml")
    
    xml_string = prettify_xml(root)
    
    # Clean up XML declaration if duplicated
    lines = xml_string.split('\n')
    if lines[0].startswith('<?xml'):
        lines = lines[1:]
    
    with open(output_path, 'w', encoding='utf-8') as f:
        f.write('<?xml version="1.0" encoding="utf-8"?>\n')
        f.write('\n'.join(lines))
    
    return str(output_path)


# =============================================================================
# CONVERSION SERVICE CLASS
# =============================================================================

class ConversionService:
    """Service layer for handling XML-Excel conversions."""
    
    def __init__(self, max_file_size_mb: int = 50):
        self.max_file_size_bytes = max_file_size_mb * 1024 * 1024
        logger.info(f"ConversionService initialized (max file size: {max_file_size_mb}MB)")
    
    def validate_file(self, file_path: str, expected_ext: Optional[str] = None) -> None:
        """Validate file existence, size, and extension."""
        path = Path(file_path)
        
        if not path.exists():
            raise ValidationError(f"File not found: {file_path}")
        
        if not path.is_file():
            raise ValidationError(f"Path is not a file: {file_path}")
        
        file_size = os.path.getsize(file_path)
        if file_size == 0:
            raise ValidationError("File is empty")
        
        if file_size > self.max_file_size_bytes:
            size_mb = file_size / (1024 * 1024)
            max_mb = self.max_file_size_bytes / (1024 * 1024)
            raise ValidationError(
                f"File too large: {size_mb:.2f}MB (max: {max_mb}MB)"
            )
        
        if expected_ext:
            if not isinstance(expected_ext, (list, tuple)):
                expected_ext = [expected_ext]
            
            if path.suffix.lower() not in [e.lower() for e in expected_ext]:
                raise ValidationError(
                    f"Invalid file extension: {path.suffix}. Expected: {', '.join(expected_ext)}"
                )
        
        logger.info(f"File validated: {file_path} ({file_size} bytes)")
    
    def xml_to_excel(self, xml_path: str, output_path: Optional[str] = None) -> str:
        """Convert XML file to Excel with validation and error handling."""
        try:
            self.validate_file(xml_path, expected_ext='.xml')
            logger.info(f"Starting XML to Excel conversion: {xml_path}")
            
            result_path = xml_to_excel_internal(xml_path, output_path)
            
            if not os.path.exists(result_path):
                raise FileProcessingError("Output file was not created")
            
            logger.info(f"Conversion successful: {result_path}")
            return result_path
            
        except ValidationError:
            raise
        except FileNotFoundError as e:
            raise ValidationError(f"File not found: {str(e)}")
        except ValueError as e:
            raise ConversionError(f"Invalid XML format: {str(e)}")
        except Exception as e:
            logger.error(f"XML to Excel conversion failed: {str(e)}", exc_info=True)
            raise ConversionError(f"Conversion failed: {str(e)}")
    
    def excel_to_xml(self, excel_path: str, output_path: Optional[str] = None, 
                     root_tag: Optional[str] = None) -> str:
        """Convert Excel file to XML with validation and error handling."""
        try:
            self.validate_file(excel_path, expected_ext=['.xlsx', '.xls', '.xlsm'])
            logger.info(f"Starting Excel to XML conversion: {excel_path}")
            
            result_path = excel_to_xml_internal(excel_path, output_path, root_tag)
            
            if not os.path.exists(result_path):
                raise FileProcessingError("Output file was not created")
            
            logger.info(f"Conversion successful: {result_path}")
            return result_path
            
        except ValidationError:
            raise
        except FileNotFoundError as e:
            raise ValidationError(f"File not found: {str(e)}")
        except ValueError as e:
            raise ConversionError(f"Invalid Excel format: {str(e)}")
        except Exception as e:
            logger.error(f"Excel to XML conversion failed: {str(e)}", exc_info=True)
            raise ConversionError(f"Conversion failed: {str(e)}")
    
    def get_file_info(self, file_path: str) -> dict:
        """Get information about a file."""
        path = Path(file_path)
        
        if not path.exists():
            raise ValidationError(f"File not found: {file_path}")
        
        file_size = os.path.getsize(file_path)
        
        return {
            "filename": path.name,
            "extension": path.suffix,
            "size_bytes": file_size,
            "size_mb": round(file_size / (1024 * 1024), 2),
            "is_xml": path.suffix.lower() == '.xml',
            "is_excel": path.suffix.lower() in ['.xlsx', '.xls', '.xlsm']
        }